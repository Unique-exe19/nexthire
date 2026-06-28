import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Recommended Candidates"

# Style header
header_font = ws.views.sheetView[0]  # Just default styling
ws.append(["candidate_id", "rank", "score", "reasoning"])

with open("submission.csv", "r", encoding="utf-8") as f:
    reader = csv.reader(f)
    # Skip header row in CSV
    next(reader)
    for row in reader:
        if not row or not row[0].strip():
            continue
        try:
            candidate_id = row[0]
            rank = int(row[1])
            score = float(row[2])
            reasoning = row[3]
            ws.append([candidate_id, rank, score, reasoning])
        except Exception as e:
            ws.append(row)

# Auto-adjust column widths
for col in ws.columns:
    max_len = 0
    for cell in col:
        if cell.value is not None:
            max_len = max(max_len, len(str(cell.value)))
    col_letter = get_column_letter(col[0].column)
    ws.column_dimensions[col_letter].width = min(max_len + 3, 100) # Cap at 100 characters

wb.save("submission.xlsx")
print("Successfully generated submission.xlsx")
